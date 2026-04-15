import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "full_ drink menu- revised.xlsx"
OUTPUT_TS = ROOT / "src" / "data" / "menuData.ts"


def _ts_string(value: str) -> str:
    escaped = (
        value.replace("\\", "\\\\")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\n", "\\n")
        .replace('"', '\\"')
    )
    return f'"{escaped}"'


def _ts_obj(d: dict[str, Any], indent: int = 0) -> str:
    pad = "  " * indent
    parts: list[str] = ["{"]
    for k, v in d.items():
        if v is None:
            continue
        if isinstance(v, str):
            parts.append(f"{pad}  {k}: {_ts_string(v)},")
        elif isinstance(v, (int, float)):
            # Prices can come as numbers in the sheet; keep them as strings in TS.
            parts.append(f"{pad}  {k}: {_ts_string(str(int(v)) if float(v).is_integer() else str(v))},")
        elif isinstance(v, list):
            parts.append(f"{pad}  {k}: [")
            for item in v:
                if isinstance(item, dict):
                    parts.append(f"{pad}    {_ts_obj(item, indent + 2)},")
                else:
                    parts.append(f"{pad}    {_ts_string(str(item))},")
            parts.append(f"{pad}  ],")
        elif isinstance(v, dict):
            parts.append(f"{pad}  {k}: {_ts_obj(v, indent + 1)},")
        else:
            parts.append(f"{pad}  {k}: {_ts_string(str(v))},")
    parts.append(f"{pad}}}")
    return "\n".join(parts)


def _slugify(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "section"


def _cell_to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        vv = v.strip()
        return vv or None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip() or None


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing xlsx: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "Full Menu" not in wb.sheetnames:
        raise SystemExit(f"Sheet 'Full Menu' not found. Found: {wb.sheetnames}")

    ws = wb["Full Menu"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h}

    required = ["Category", "Item", "Description", "DOM", "IMP", "Bottle"]
    missing = [k for k in required if k not in idx]
    if missing:
        raise SystemExit(f"Missing columns in header: {missing}. Found: {header}")

    rows: list[dict[str, str | None]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        cat = _cell_to_str(r[idx["Category"]])
        item = _cell_to_str(r[idx["Item"]])
        if not cat or not item:
            continue
        rows.append(
            {
                "category": cat,
                "item": item,
                "description": _cell_to_str(r[idx["Description"]]),
                "dom": _cell_to_str(r[idx["DOM"]]),
                "imp": _cell_to_str(r[idx["IMP"]]),
                "bottle": _cell_to_str(r[idx["Bottle"]]),
            }
        )

    def mk_item(row: dict[str, str | None]) -> dict[str, Any]:
        d: dict[str, Any] = {"name": row["item"]}
        if row.get("description"):
            d["description"] = row["description"]
        return d

    # Simple one-column price categories (use DOM as price when present, else IMP, else Bottle)
    def mk_single_price_item(row: dict[str, str | None]) -> dict[str, Any]:
        d = mk_item(row)
        price = row.get("dom") or row.get("imp") or row.get("bottle")
        if price:
            d["price"] = price
        return d

    # Dual DOM/IMP price items
    def mk_dom_imp_item(row: dict[str, str | None]) -> dict[str, Any]:
        d = mk_item(row)
        if row.get("dom"):
            d["priceDom"] = row["dom"]
        if row.get("imp"):
            d["priceImp"] = row["imp"]
        return d

    # Glass/Bottle style (used for Wine and spirits PEG/Bottle)
    def mk_glass_bottle_item(row: dict[str, str | None]) -> dict[str, Any]:
        d = mk_item(row)
        glass = row.get("dom") or row.get("imp")
        bottle = row.get("bottle")
        if glass:
            d["priceGlass"] = glass
        if bottle:
            d["priceBottle"] = bottle
        return d

    # Beer: domestic (300/650), imported (single price), draught (single price via Bottle col)
    beer_imported: list[dict[str, Any]] = []
    beer_domestic: list[dict[str, Any]] = []
    beer_draught: list[dict[str, Any]] = []

    wine_items: list[dict[str, Any]] = []

    cocktails: list[dict[str, Any]] = []
    shooters: list[dict[str, Any]] = []

    pitchers: list[dict[str, Any]] = []
    mixers: list[dict[str, Any]] = []
    mocktails: list[dict[str, Any]] = []
    milkshakes: list[dict[str, Any]] = []
    cold_bev: list[dict[str, Any]] = []
    hot_bev: list[dict[str, Any]] = []

    liquor: list[dict[str, Any]] = []
    vodka: list[dict[str, Any]] = []
    gin: list[dict[str, Any]] = []
    tequila: list[dict[str, Any]] = []
    rum: list[dict[str, Any]] = []
    brandy: list[dict[str, Any]] = []

    whiskey_single: list[dict[str, Any]] = []
    whiskey_blended: list[dict[str, Any]] = []
    whiskey_bourbon: list[dict[str, Any]] = []
    whiskey_domestic: list[dict[str, Any]] = []

    for row in rows:
        cat = row["category"]
        if cat == "Cocktails":
            cocktails.append(mk_dom_imp_item(row))
        elif cat == "Shooters":
            shooters.append(mk_dom_imp_item(row))
        elif cat == "Pitchers":
            pitchers.append(mk_single_price_item(row))
        elif cat == "Mixers":
            mixers.append(mk_single_price_item(row))
        elif cat == "Mocktails":
            mocktails.append(mk_single_price_item(row))
        elif cat == "Milkshakes":
            milkshakes.append(mk_single_price_item(row))
        elif cat == "Cold Beverages":
            cold_bev.append(mk_single_price_item(row))
        elif cat == "Hot Beverages":
            hot_bev.append(mk_single_price_item(row))
        elif cat == "Liquor":
            liquor.append(mk_glass_bottle_item(row))
        elif cat == "Vodka":
            vodka.append(mk_glass_bottle_item(row))
        elif cat == "Gin":
            gin.append(mk_glass_bottle_item(row))
        elif cat == "Tequila":
            tequila.append(mk_glass_bottle_item(row))
        elif cat == "Rum":
            rum.append(mk_glass_bottle_item(row))
        elif cat == "Brandy":
            brandy.append(mk_glass_bottle_item(row))
        elif cat == "Wine":
            wine_items.append(mk_glass_bottle_item(row))
        elif cat == "Beer":
            name = row["item"] or ""
            if (row.get("bottle") and not row.get("dom") and not row.get("imp")) or "draught" in name.lower():
                beer_draught.append(mk_single_price_item({"item": name, "description": row.get("description"), "dom": None, "imp": None, "bottle": row.get("bottle"), "category": cat}))
            elif row.get("dom"):
                d = mk_item(row)
                d["price300"] = row.get("dom")
                if row.get("imp"):
                    d["price650"] = row.get("imp")
                beer_domestic.append(d)
            else:
                beer_imported.append(mk_single_price_item(row))
        elif cat == "Single Malt Whiskey":
            whiskey_single.append(mk_glass_bottle_item(row))
        elif cat == "Blended Scotch":
            whiskey_blended.append(mk_glass_bottle_item(row))
        elif cat == "Bourbon Whiskey":
            whiskey_bourbon.append(mk_glass_bottle_item(row))
        elif cat == "Domestic Whiskey":
            # Domestic whiskey uses DOM for peg in the sheet
            whiskey_domestic.append(mk_glass_bottle_item(row))
        else:
            # Unknown category; ignore for now.
            pass

    menu_sections: list[dict[str, Any]] = [
        {"id": "cocktails", "title": "Cocktails", "priceHeaders": ["DOM", "IMP"], "items": cocktails},
        {"id": "mocktails", "title": "Mocktails", "items": mocktails},
        {"id": "milkshakes", "title": "Milkshakes", "items": milkshakes},
        {"id": "hot-beverages", "title": "Hot Beverages", "items": hot_bev},
        {"id": "cold-beverages", "title": "Cold Beverages", "items": cold_bev},
        {"id": "pitchers", "title": "Pitchers", "items": pitchers},
        {"id": "shooters", "title": "Shooters", "priceHeaders": ["DOM", "IMP"], "items": shooters},
        {
            "id": "whiskey",
            "title": "Whiskey",
            "items": [],
            "priceHeaders": ["Peg", "Bottle"],
            "subsections": [
                {"title": "Single Malt Whiskey", "items": whiskey_single},
                {"title": "Blended Scotch Whiskey", "items": whiskey_blended},
                {"title": "Bourbon Whiskey", "items": whiskey_bourbon},
                {"title": "Domestic Whiskey", "items": whiskey_domestic},
            ],
        },
        {"id": "vodka", "title": "Vodka", "priceHeaders": ["Peg", "Bottle"], "items": vodka},
        {"id": "gin", "title": "Gin", "priceHeaders": ["Peg", "Bottle"], "items": gin},
        {"id": "tequila", "title": "Tequila", "priceHeaders": ["Peg", "Bottle"], "items": tequila},
        {"id": "rum", "title": "Rum", "priceHeaders": ["Peg", "Bottle"], "items": rum},
        {"id": "brandy", "title": "Brandy", "priceHeaders": ["Peg", "Bottle"], "items": brandy},
        {"id": "liqueur", "title": "Liqueur", "priceHeaders": ["Peg", "Bottle"], "items": liquor},
        {"id": "wine", "title": "Wine", "priceHeaders": ["Glass", "Bottle"], "items": wine_items},
        {
            "id": "beer",
            "title": "Beer",
            "items": [],
            "subsections": [
                {"title": "Imported", "items": beer_imported},
                {"title": "Domestic", "items": beer_domestic},
                {"title": "Draught", "items": beer_draught},
            ],
        },
        {"id": "mixers", "title": "Mixers", "items": mixers},
    ]

    sections_ts = ",\n".join("  " + _ts_obj(sec, 1).replace("\n", "\n  ") for sec in menu_sections)

    ts = f'''export interface MenuItem {{
  name: string;
  description?: string;
  price?: string;
  priceDom?: string;
  priceImp?: string;
  priceGlass?: string;
  priceBottle?: string;
  price300?: string;
  price650?: string;
}}

export interface MenuSection {{
  id: string;
  title: string;
  subtitle?: string;
  priceHeaders?: [string, string];
  items: MenuItem[];
  subsections?: {{ title: string; items: MenuItem[] }}[];
}}

export const menuSections: MenuSection[] = [
{sections_ts}
];
'''

    OUTPUT_TS.write_text(ts, encoding="utf-8")
    print(f"Wrote {OUTPUT_TS}")


if __name__ == "__main__":
    main()

